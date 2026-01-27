#!/usr/bin/env python3
"""
Compare v6 Baseline and v8 Hybrid methodology forecasts against budget targets.
Shows whether v8 improvements close the gap relative to v6 baseline.
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import warnings
warnings.filterwarnings("ignore")

BUDGET_FILE = "Budget consol file.xlsx"
V6_FILE = "output_v6_baseline_full/Forecast_Transparency_Report.xlsx"
V8_FILE = "output_v8_hybrid_full/Forecast_Transparency_Report.xlsx"

BUDGET_ROWS = {
    "Collections":      {"Non Prime": 12, "NPS": 13, "NPM": 14, "Prime": 15, "Total": 16},
    "ClosingGBV":       {"Non Prime": 23, "NPS": 24, "NPM": 25, "Prime": 26, "Total": 27},
    "ClosingNBV":       {"Non Prime": 43, "NPS": 44, "NPM": 45, "Prime": 46, "Total": 47},
    "Revenue":          {"Non Prime": 63, "NPS": 64, "NPM": 65, "Prime": 66, "Total": 67},
    "GrossImpairment":  {"Non Prime": 74, "NPS": 75, "NPM": 76, "Prime": 77, "Total": 78},
    "NetImpairment":    {"Non Prime": 122, "NPS": 123, "NPM": 124, "Prime": 125, "Total": 126},
}

FORECAST_SEG_MAP = {
    "Non Prime": ["NON PRIME"],
    "NPS": ["NRP-S"],
    "NPM": ["NRP-M", "NRP-L"],
    "Prime": ["PRIME"],
}


def load_budget():
    wb = openpyxl.load_workbook(BUDGET_FILE, data_only=True)
    ws = wb["P&L analysis - BB"]
    month_cols = {}
    for col in range(4, ws.max_column + 1):
        dv = ws.cell(row=3, column=col).value
        if dv is None:
            continue
        if isinstance(dv, (int, float)):
            dv = datetime(1899, 12, 30) + timedelta(days=int(dv))
        if isinstance(dv, datetime):
            month_cols[dv.strftime("%Y-%m")] = col
    budget = {}
    for metric, rows in BUDGET_ROWS.items():
        budget[metric] = {}
        for seg, rn in rows.items():
            budget[metric][seg] = {}
            for mk, col in month_cols.items():
                val = ws.cell(row=rn, column=col).value
                if val is not None:
                    try:
                        budget[metric][seg][mk] = float(val)
                    except:
                        pass
    wb.close()
    return budget


def load_forecast(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    def read_sheet(name):
        ws = wb[name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for c, h in enumerate(headers, 1):
                if h:
                    row[h] = ws.cell(row=r, column=c).value
            if row.get("Segment"):
                rows.append(row)
        return rows

    summary = read_sheet("9_Summary")
    impairment = read_sheet("11_Impairment")
    wb.close()

    forecast = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for row in summary:
        seg = row.get("Segment")
        mv = row.get("ForecastMonth") or row.get("CalendarMonth")
        if not seg or not mv:
            continue
        if isinstance(mv, datetime):
            mk = mv.strftime("%Y-%m")
        elif isinstance(mv, str):
            try:
                mk = datetime.strptime(mv, "%Y-%m-%d").strftime("%Y-%m")
            except:
                continue
        else:
            continue

        for bseg, fsegs in FORECAST_SEG_MAP.items():
            if seg in fsegs:
                coll_p = float(row.get("Coll_Principal", 0) or 0)
                coll_i = float(row.get("Coll_Interest", 0) or 0)
                forecast["Collections"][bseg][mk] += abs(coll_p) + abs(coll_i)
                forecast["ClosingGBV"][bseg][mk] += float(row.get("ClosingGBV", 0) or 0)
                forecast["Revenue"][bseg][mk] += float(row.get("InterestRevenue", 0) or 0)
                break

    for row in impairment:
        seg = row.get("Segment")
        mv = row.get("ForecastMonth") or row.get("CalendarMonth")
        if not seg or not mv:
            continue
        if isinstance(mv, datetime):
            mk = mv.strftime("%Y-%m")
        elif isinstance(mv, str):
            try:
                mk = datetime.strptime(mv, "%Y-%m-%d").strftime("%Y-%m")
            except:
                continue
        else:
            continue

        for bseg, fsegs in FORECAST_SEG_MAP.items():
            if seg in fsegs:
                forecast["ClosingNBV"][bseg][mk] += float(row.get("ClosingNBV", 0) or 0)
                gi = float(row.get("Gross_Impairment_ExcludingDS", 0) or 0)
                ni = float(row.get("Net_Impairment", 0) or 0)
                forecast["GrossImpairment"][bseg][mk] += gi
                forecast["NetImpairment"][bseg][mk] += ni
                break

    for metric in forecast:
        all_months = set()
        for s in FORECAST_SEG_MAP:
            all_months.update(forecast[metric][s].keys())
        for mk in all_months:
            forecast[metric]["Total"][mk] = sum(
                forecast[metric][s].get(mk, 0) for s in FORECAST_SEG_MAP
            )

    return forecast


def var_pct(f, b):
    if b == 0:
        return float('nan')
    return (f - b) / abs(b) * 100


def main():
    print("=" * 120)
    print("  METHODOLOGY COMPARISON: v6 Baseline vs v8 Hybrid vs Budget")
    print("=" * 120)

    budget = load_budget()
    v6 = load_forecast(V6_FILE)
    v8 = load_forecast(V8_FILE)

    # Find overlapping months
    bm = set()
    for m in budget:
        for s in budget[m]:
            bm.update(budget[m][s].keys())
    fm = set()
    for m in v6:
        for s in v6[m]:
            fm.update(v6[m][s].keys())
    overlap = sorted(bm & fm)
    print(f"\n  Overlapping months: {overlap[0]} to {overlap[-1]} ({len(overlap)} months)")

    metrics = ["Collections", "ClosingGBV", "ClosingNBV", "Revenue", "GrossImpairment", "NetImpairment"]
    stock_metrics = {"ClosingGBV", "ClosingNBV"}

    # ==========================================
    # TOTAL PORTFOLIO SUMMARY
    # ==========================================
    print(f"\n{'='*120}")
    print(f"  TOTAL PORTFOLIO SUMMARY (cumulative over {len(overlap)} months)")
    print(f"{'='*120}")
    print(f"  {'Metric':<22} {'Budget':>14} {'v6 Baseline':>14} {'v6 Var%':>10} {'v8 Hybrid':>14} {'v8 Var%':>10} {'Improvement':>14}")
    print(f"  {'-'*100}")

    for metric in metrics:
        is_stock = metric in stock_metrics
        b_total = 0
        v6_total = 0
        v8_total = 0

        for m in overlap:
            bv = budget.get(metric, {}).get("Total", {}).get(m, None)
            f6 = v6.get(metric, {}).get("Total", {}).get(m, 0)
            f8 = v8.get(metric, {}).get("Total", {}).get(m, 0)
            if bv is not None:
                if is_stock:
                    b_total = bv
                    v6_total = f6
                    v8_total = f8
                else:
                    b_total += bv
                    v6_total += f6
                    v8_total += f8

        if b_total != 0:
            v6_pct = var_pct(v6_total, b_total)
            v8_pct = var_pct(v8_total, b_total)
            improvement = abs(v6_pct) - abs(v8_pct)
            imp_str = f"+{improvement:.1f}pp" if improvement > 0 else f"{improvement:.1f}pp"
            print(f"  {metric:<22} {b_total:>14,.0f} {v6_total:>14,.0f} {v6_pct:>+9.1f}% {v8_total:>14,.0f} {v8_pct:>+9.1f}% {imp_str:>14}")

    # ==========================================
    # BY SEGMENT DETAIL
    # ==========================================
    seg_labels = {
        "Non Prime": "Non Prime",
        "NPS":       "NRP-S    ",
        "NPM":       "NRP-M+L  ",
        "Prime":     "Prime    ",
        "Total":     "TOTAL    ",
    }
    segments = ["Non Prime", "NPS", "NPM", "Prime", "Total"]

    for metric in metrics:
        is_stock = metric in stock_metrics
        print(f"\n{'='*120}")
        print(f"  {metric} by Segment")
        print(f"{'='*120}")
        print(f"  {'Segment':<14} {'Budget':>14} {'v6 Baseline':>14} {'v6 Var%':>10} {'v8 Hybrid':>14} {'v8 Var%':>10} {'Gap Closed':>12}")
        print(f"  {'-'*80}")

        for seg in segments:
            b_s = 0
            v6_s = 0
            v8_s = 0
            for m in overlap:
                bv = budget.get(metric, {}).get(seg, {}).get(m, None)
                f6 = v6.get(metric, {}).get(seg, {}).get(m, 0)
                f8 = v8.get(metric, {}).get(seg, {}).get(m, 0)
                if bv is not None:
                    if is_stock:
                        b_s = bv
                        v6_s = f6
                        v8_s = f8
                    else:
                        b_s += bv
                        v6_s += f6
                        v8_s += f8

            if b_s != 0:
                v6p = var_pct(v6_s, b_s)
                v8p = var_pct(v8_s, b_s)
                improvement = abs(v6p) - abs(v8p)
                imp_str = f"+{improvement:.1f}pp" if improvement > 0 else f"{improvement:.1f}pp"
                print(f"  {seg_labels[seg]:<14} {b_s:>14,.0f} {v6_s:>14,.0f} {v6p:>+9.1f}% {v8_s:>14,.0f} {v8p:>+9.1f}% {imp_str:>12}")

    # ==========================================
    # MONTH-BY-MONTH FOR KEY METRICS
    # ==========================================
    print(f"\n{'='*120}")
    print(f"  MONTH-BY-MONTH: Total Portfolio")
    print(f"{'='*120}")

    for metric in ["Collections", "ClosingGBV", "GrossImpairment", "ClosingNBV"]:
        is_stock = metric in stock_metrics
        print(f"\n  {metric}:")
        print(f"  {'Month':<10} {'Budget':>14} {'v6':>14} {'v6%':>8} {'v8':>14} {'v8%':>8}")
        print(f"  {'-'*72}")
        for m in overlap:
            bv = budget.get(metric, {}).get("Total", {}).get(m, 0)
            f6 = v6.get(metric, {}).get("Total", {}).get(m, 0)
            f8 = v8.get(metric, {}).get("Total", {}).get(m, 0)
            if bv:
                v6p = var_pct(f6, bv)
                v8p = var_pct(f8, bv)
                print(f"  {m:<10} {bv:>14,.0f} {f6:>14,.0f} {v6p:>+7.1f}% {f8:>14,.0f} {v8p:>+7.1f}%")

    print(f"\n{'='*120}")
    print(f"  COVERAGE RATIO COMPARISON (forecast vs budget implied)")
    print(f"{'='*120}")
    for seg in ["Non Prime", "NPS", "NPM", "Prime", "Total"]:
        print(f"\n  {seg_labels[seg].strip()}:")
        print(f"  {'Month':<10} {'Budget CR':>12} {'v6 CR':>12} {'v8 CR':>12}")
        print(f"  {'-'*50}")
        for m in overlap:
            b_gbv = budget.get("ClosingGBV", {}).get(seg, {}).get(m, 0)
            b_nbv = budget.get("ClosingNBV", {}).get(seg, {}).get(m, 0)
            b_cr = (b_gbv - b_nbv) / b_gbv * 100 if b_gbv else 0

            v6_gbv = v6.get("ClosingGBV", {}).get(seg, {}).get(m, 0)
            v6_nbv = v6.get("ClosingNBV", {}).get(seg, {}).get(m, 0)
            v6_cr = (v6_gbv - v6_nbv) / v6_gbv * 100 if v6_gbv else 0

            v8_gbv = v8.get("ClosingGBV", {}).get(seg, {}).get(m, 0)
            v8_nbv = v8.get("ClosingNBV", {}).get(seg, {}).get(m, 0)
            v8_cr = (v8_gbv - v8_nbv) / v8_gbv * 100 if v8_gbv else 0

            print(f"  {m:<10} {b_cr:>11.2f}% {v6_cr:>11.2f}% {v8_cr:>11.2f}%")


if __name__ == "__main__":
    main()
