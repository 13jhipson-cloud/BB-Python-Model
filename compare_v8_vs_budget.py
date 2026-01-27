#!/usr/bin/env python3
"""
Compare v8 Hybrid methodology forecast output against budget targets.
Reads from:
  - output_v8_hybrid/Forecast_Transparency_Report.xlsx (9_Summary and 11_Impairment sheets)
  - Budget consol file.xlsx (P&L analysis - BB sheet)
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import warnings
warnings.filterwarnings("ignore")

BUDGET_FILE = "Budget consol file.xlsx"
FORECAST_FILE = "output_v8_hybrid/Forecast_Transparency_Report.xlsx"

BUDGET_ROWS = {
    "Collections":      {"Non Prime": 12, "NPS": 13, "NPM": 14, "Prime": 15, "Total": 16},
    "ClosingGBV":       {"Non Prime": 23, "NPS": 24, "NPM": 25, "Prime": 26, "Total": 27},
    "ClosingNBV":       {"Non Prime": 43, "NPS": 44, "NPM": 45, "Prime": 46, "Total": 47},
    "Revenue":          {"Non Prime": 63, "NPS": 64, "NPM": 65, "Prime": 66, "Total": 67},
    "GrossImpairment":  {"Non Prime": 74, "NPS": 75, "NPM": 76, "Prime": 77, "Total": 78},
    "NetImpairment":    {"Non Prime": 122, "NPS": 123, "NPM": 124, "Prime": 125, "Total": 126},
}

FORECAST_SEG_MAP = {
    "Non Prime": ["NON PRIME"],
    "NPS": ["NRP-S"],
    "NPM": ["NRP-M", "NRP-L"],
    "Prime": ["PRIME"],
}


def load_budget_data():
    wb = openpyxl.load_workbook(BUDGET_FILE, data_only=True)
    ws = wb["P&L analysis - BB"]
    month_cols = {}
    for col in range(4, ws.max_column + 1):
        date_val = ws.cell(row=3, column=col).value
        if date_val is None:
            continue
        if isinstance(date_val, (int, float)):
            base = datetime(1899, 12, 30)
            date_val = base + timedelta(days=int(date_val))
        if isinstance(date_val, datetime):
            key = date_val.strftime("%Y-%m")
            month_cols[key] = col

    budget = {}
    for metric, rows in BUDGET_ROWS.items():
        budget[metric] = {}
        for seg, row_num in rows.items():
            budget[metric][seg] = {}
            for month_key, col in month_cols.items():
                val = ws.cell(row=row_num, column=col).value
                if val is not None:
                    try:
                        budget[metric][seg][month_key] = float(val)
                    except (ValueError, TypeError):
                        pass
    wb.close()
    return budget


def load_forecast_data():
    wb = openpyxl.load_workbook(FORECAST_FILE, data_only=True)

    # Load summary (9_Summary)
    ws = wb["9_Summary"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            if h:
                row[h] = ws.cell(row=r, column=c).value
        if row.get("Segment"):
            rows.append(row)

    # Load impairment (11_Impairment)
    ws_imp = wb["11_Impairment"]
    imp_headers = [ws_imp.cell(row=1, column=c).value for c in range(1, ws_imp.max_column + 1)]
    imp_rows = []
    for r in range(2, ws_imp.max_row + 1):
        row = {}
        for c, h in enumerate(imp_headers, 1):
            if h:
                row[h] = ws_imp.cell(row=r, column=c).value
        if row.get("Segment"):
            imp_rows.append(row)

    wb.close()
    return rows, imp_rows


def aggregate_forecast(summary_rows, imp_rows):
    """Aggregate forecast data by budget segment mapping."""
    forecast = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for row in summary_rows:
        seg = row.get("Segment")
        month_val = row.get("ForecastMonth") or row.get("CalendarMonth")
        if not seg or not month_val:
            continue
        if isinstance(month_val, datetime):
            month_key = month_val.strftime("%Y-%m")
        elif isinstance(month_val, str):
            try:
                month_key = datetime.strptime(month_val, "%Y-%m-%d").strftime("%Y-%m")
            except:
                continue
        else:
            continue

        for bseg, fsegs in FORECAST_SEG_MAP.items():
            if seg in fsegs:
                forecast["Collections"][bseg][month_key] += float(row.get("Coll_Principal", 0) or 0) + float(row.get("Coll_Interest", 0) or 0)
                forecast["ClosingGBV"][bseg][month_key] += float(row.get("ClosingGBV", 0) or 0)
                forecast["Revenue"][bseg][month_key] += float(row.get("InterestRevenue", 0) or 0)
                break

    for row in imp_rows:
        seg = row.get("Segment")
        month_val = row.get("ForecastMonth") or row.get("CalendarMonth")
        if not seg or not month_val:
            continue
        if isinstance(month_val, datetime):
            month_key = month_val.strftime("%Y-%m")
        elif isinstance(month_val, str):
            try:
                month_key = datetime.strptime(month_val, "%Y-%m-%d").strftime("%Y-%m")
            except:
                continue
        else:
            continue

        for bseg, fsegs in FORECAST_SEG_MAP.items():
            if seg in fsegs:
                forecast["ClosingNBV"][bseg][month_key] += float(row.get("ClosingNBV", 0) or 0)
                forecast["GrossImpairment"][bseg][month_key] += float(row.get("Gross_Impairment_ExcludingDS", 0) or 0)
                forecast["NetImpairment"][bseg][month_key] += float(row.get("Net_Impairment", 0) or 0)
                break

    # Compute Total across segments
    for metric in forecast:
        for month_key in set().union(*[forecast[metric][s].keys() for s in FORECAST_SEG_MAP]):
            forecast[metric]["Total"][month_key] = sum(
                forecast[metric][s].get(month_key, 0) for s in FORECAST_SEG_MAP
            )

    return forecast


def main():
    print("=" * 100)
    print("  V8 HYBRID METHODOLOGY: FORECAST vs BUDGET COMPARISON")
    print("=" * 100)

    budget = load_budget_data()
    summary_rows, imp_rows = load_forecast_data()
    forecast = aggregate_forecast(summary_rows, imp_rows)

    # Find overlapping months
    budget_months = set()
    for metric in budget:
        for seg in budget[metric]:
            budget_months.update(budget[metric][seg].keys())

    forecast_months = set()
    for metric in forecast:
        for seg in forecast[metric]:
            forecast_months.update(forecast[metric][seg].keys())

    overlap = sorted(budget_months & forecast_months)
    print(f"\n  Overlapping months: {overlap[0]} to {overlap[-1]} ({len(overlap)} months)")

    # Compare by metric and segment
    metrics = ["Collections", "ClosingGBV", "ClosingNBV", "Revenue", "GrossImpairment", "NetImpairment"]
    segments = ["Non Prime", "NPS", "NPM", "Prime", "Total"]
    seg_labels = {
        "Non Prime": "Non Prime    ",
        "NPS":       "NRP-S        ",
        "NPM":       "NRP-M+L      ",
        "Prime":     "Prime        ",
        "Total":     "TOTAL        ",
    }

    for metric in metrics:
        print(f"\n{'='*100}")
        print(f"  {metric}")
        print(f"{'='*100}")

        is_stock = metric in ("ClosingGBV", "ClosingNBV")

        for seg in segments:
            b_sum = 0
            f_sum = 0
            count = 0
            for m in overlap:
                b_val = budget.get(metric, {}).get(seg, {}).get(m, None)
                f_val = forecast.get(metric, {}).get(seg, {}).get(m, 0)
                if b_val is not None:
                    if is_stock:
                        b_sum = b_val  # Last value for stock metrics
                        f_sum = f_val
                    else:
                        b_sum += b_val
                        f_sum += f_val
                    count += 1

            if count > 0 and b_sum != 0:
                var_pct = (f_sum - b_sum) / abs(b_sum) * 100
                var_abs = f_sum - b_sum
                label = "Avg" if is_stock else "Sum"
                print(f"  {seg_labels[seg]}  Budget: {b_sum:>14,.0f}  Forecast: {f_sum:>14,.0f}  Var: {var_abs:>+12,.0f} ({var_pct:>+7.1f}%)  [{label} over {count} months]")
            elif count > 0:
                print(f"  {seg_labels[seg]}  Budget: {'N/A':>14}  Forecast: {f_sum:>14,.0f}")

    # Summary table
    print(f"\n{'='*100}")
    print(f"  SUMMARY: Total Portfolio Variance (v8 Hybrid vs Budget)")
    print(f"{'='*100}")
    print(f"  {'Metric':<25} {'Budget':>14} {'Forecast':>14} {'Variance':>12} {'Var %':>8}")
    print(f"  {'-'*75}")

    for metric in metrics:
        is_stock = metric in ("ClosingGBV", "ClosingNBV")
        b_sum = 0
        f_sum = 0
        for m in overlap:
            b_val = budget.get(metric, {}).get("Total", {}).get(m, None)
            f_val = forecast.get(metric, {}).get("Total", {}).get(m, 0)
            if b_val is not None:
                if is_stock:
                    b_sum = b_val
                    f_sum = f_val
                else:
                    b_sum += b_val
                    f_sum += f_val

        if b_sum != 0:
            var_pct = (f_sum - b_sum) / abs(b_sum) * 100
            var_abs = f_sum - b_sum
            print(f"  {metric:<25} {b_sum:>14,.0f} {f_sum:>14,.0f} {var_abs:>+12,.0f} {var_pct:>+7.1f}%")

    # Compare with previous iterations
    print(f"\n{'='*100}")
    print(f"  COMPARISON WITH PREVIOUS CALIBRATION ITERATIONS")
    print(f"{'='*100}")
    print(f"  {'Iteration':<30} {'Collections':>12} {'GBV':>12} {'NBV':>12} {'GrossImp':>12}")
    print(f"  {'-'*80}")
    print(f"  {'Iter 0 (Baseline v6)':30} {'+10.1%':>12} {'-9.0%':>12} {'-3.4%':>12} {'-57.8%':>12}")
    print(f"  {'Iter 1 (ScaledCA 1.8x)':30} {'N/A':>12} {'N/A':>12} {'-29.0%':>12} {'+166.0%':>12}")
    print(f"  {'Iter 2 (ScaledCA 1.2x)':30} {'N/A':>12} {'N/A':>12} {'-3.9%':>12} {'-79.7%':>12}")
    print(f"  {'V8 HYBRID (this run)':30} {'See above':>12} {'See above':>12} {'See above':>12} {'See above':>12}")
    print()


if __name__ == "__main__":
    main()
